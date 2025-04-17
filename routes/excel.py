from flask import Blueprint, send_file
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import io
from datetime import datetime
from models.database import get_db_connection
import os
from openpyxl.drawing.image import Image

excel_bp = Blueprint('excel', __name__)

# 한글 금액 변환
def number_to_korean(num):
    units = ["", "만", "억", "조"]
    result = ""
    split_unit = 10000
    i = 0
    while num > 0:
        part = num % split_unit
        if part:
            result = str(part) + units[i] + result
        num //= split_unit
        i += 1
    return "일금" + result + "원정 (VAT포함)"

# 병합 셀 안전하게 쓰기
def write_to_merged_auto(ws, target_cell, value):
    for merged_range in ws.merged_cells.ranges:
        if target_cell in merged_range:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            ws.unmerge_cells(str(merged_range))
            ws.cell(row=min_row, column=min_col, value=value)
            ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
            return
    ws[target_cell] = value

def write_and_merge(ws, start_cell, end_cell, value):
    ws.merge_cells(f"{start_cell}:{end_cell}")
    ws[start_cell] = value

@excel_bp.route('/api/export_excel/<int:estimate_id>', methods=['GET'])
def export_estimate_excel(estimate_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    # 1. 견적서 및 고객, 담당자 정보 조회
    sql = """
    SELECT 
        e.*,  
        c.customer_nm,
        u.name, u.position, u.phone, u.email
    FROM estimate e
    LEFT JOIN customer c ON e.customer_id = c.customer_id
    LEFT JOIN user u ON e.sales_id = u.usr_id
    WHERE e.id = %s
    """
    cursor.execute(sql, (estimate_id,))
    estimate = cursor.fetchone()
    if not estimate:
        return {'error': '견적서를 찾을 수 없습니다.'}, 404

    # 2. 제품 목록 조회
    sql_products = """
    SELECT 
        p.p_name,
        p.p_description,
        p.p_price,
        ep.quantity,
        ep.unit_price,
        ep.total_price
    FROM t_estimate_product ep
    JOIN t_product_add p ON ep.product_id = p.id
    WHERE ep.estimate_id = %s
    """
    cursor.execute(sql_products, (estimate_id,))
    products = cursor.fetchall()

    # 3. 참조자 조회
    sql_reference = """
    SELECT 
        r.manager_name
    FROM estimate_reference r
    WHERE r.estimate_id = %s
    """
    cursor.execute(sql_reference, (estimate_id,))
    references = cursor.fetchall()

    cursor.close()
    conn.close()

    # 4. 엑셀 템플릿 불러오기
    EXCEL_TEMPLATE_DIR = os.path.join(os.getcwd(), "templates")
    template_path = os.path.join(EXCEL_TEMPLATE_DIR, "estimate_execl_template.xlsx")
    wb = load_workbook(template_path)
    ws = wb.active

    # 5. 상단 정보 입력
    write_to_merged_auto(ws, "B6", "(주) " + estimate["customer_nm"])
    write_to_merged_auto(ws, "H5", datetime.today().strftime("%Y년 %m월 %d일"))
    reference_names = ", ".join([ref["manager_name"] for ref in references])
    write_to_merged_auto(ws, "C7", reference_names + " 님")
    write_to_merged_auto(ws, "C8", estimate["quote_title"])
    write_to_merged_auto(ws, "B11", number_to_korean(int(estimate["total_price_with_vat"])))

    # 6. 제품 리스트 입력 (최대 30줄)
    start_row = 13
    max_product_rows = 30
    actual_product_rows = len(products)

    for idx, item in enumerate(products):
        row = start_row + idx
        ws.cell(row=row, column=2, value=idx + 1)
        ws.cell(row=row, column=3, value=item["p_name"])
        ws.cell(row=row, column=4, value=item["p_description"])
        ws.cell(row=row, column=5, value=item["quantity"])
        ws.cell(row=row, column=6, value=item["p_price"])
        ws.cell(row=row, column=7, value=item["unit_price"])
        ws.cell(row=row, column=8, value=item["total_price"])

    if actual_product_rows < max_product_rows:
        ws.delete_rows(start_row + actual_product_rows, max_product_rows - actual_product_rows)

    # 7. 합계 정보 삽입
    summary_start_row = start_row + len(products)

    write_and_merge(ws, f"B{summary_start_row}", f"F{summary_start_row}", "합        계")
    write_and_merge(ws, f"G{summary_start_row}", f"H{summary_start_row}", "₩" + str(estimate["total_price_before_vat"])) # 합        계

    write_and_merge(ws, f"B{summary_start_row + 1}", f"F{summary_start_row + 1}", "부   가   세")
    write_and_merge(ws, f"G{summary_start_row + 1}", f"H{summary_start_row + 1}", "₩" + str(estimate["vat"])) # 부   가   세

    write_and_merge(ws, f"B{summary_start_row + 2}", f"F{summary_start_row + 2}", "총   합   계 (VAT포함)")
    write_and_merge(ws, f"G{summary_start_row + 2}", f"H{summary_start_row + 2}", "₩" + str(estimate["total_price_with_vat"])) # 총   합   계 (VAT포함)

    # 하단 정보 입력
    write_and_merge(ws, f"E{summary_start_row + 3}", f"H{summary_start_row + 3}", "구매자확인")
    write_and_merge(ws, f"E{summary_start_row + 4}", f"H{summary_start_row + 4}", "당사는 이 견적서상의 가격 및 조건들을 수용하고 이 견적서를 발주서로 대신합니다.")
    
    write_and_merge(ws, f"E{summary_start_row + 5}", f"F{summary_start_row + 5}", "회사명")
    write_and_merge(ws, f"G{summary_start_row + 5}", f"H{summary_start_row + 5}", "")

    write_and_merge(ws, f"E{summary_start_row + 6}", f"F{summary_start_row + 6}", "발주담당자")
    write_and_merge(ws, f"G{summary_start_row + 6}", f"H{summary_start_row + 6}", "")

    write_and_merge(ws, f"E{summary_start_row + 7}", f"F{summary_start_row + 7}", "배송주소지")
    write_and_merge(ws, f"G{summary_start_row + 7}", f"H{summary_start_row + 7}", "")
    write_and_merge(ws, f"E{summary_start_row + 8}", f"F{summary_start_row + 8}", "대표 / 신청인")
    write_and_merge(ws, f"G{summary_start_row + 8}", f"H{summary_start_row + 8}", "                /                 (인)")



    write_and_merge(ws, f"C{summary_start_row + 3}", f"D{summary_start_row + 3}", estimate["valid_until"]) # 견적유효기간
    write_and_merge(ws, f"C{summary_start_row + 4}", f"D{summary_start_row + 4}", estimate["delivery_condition"]) # 납기
    write_and_merge(ws, f"C{summary_start_row + 5}", f"D{summary_start_row + 5}", estimate["payment_condition"]) # 결제조건
    write_and_merge(ws, f"C{summary_start_row + 6}", f"D{summary_start_row + 6}", estimate["payment_condition"]) # 하자보증기간
    write_and_merge(ws, f"C{summary_start_row + 7}", f"D{summary_start_row + 7}", f"{estimate['name']} / {estimate['position']} / {estimate['phone']} / {estimate['email']}") # 영업담당
    write_and_merge(ws, f"C{summary_start_row + 8}", f"D{summary_start_row + 8}", estimate["remarks"]) # 특이사항



    # 회사 로고 삽입 -> 



    # 9. 인쇄 영역 자동 설정 (마지막 제품 + 7줄까지)
    print_end_row = start_row + actual_product_rows + 9
    ws.print_area = f"A1:H{print_end_row}"


    # 8. 메모리에 저장 후 반환
    # output = io.BytesIO()
    # wb.save(output)
    # output.seek(0)

    # 🔥 실제 파일 경로로 저장
    filename = f"estimate_{estimate['quote_id']}_{datetime.today().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = os.path.join("/usr/local/flask/yeji/groupware-api/temp", filename)
    wb.save(file_path)

    return send_file(
        #output,
        file_path,
        as_attachment=True,
        download_name = f"견적서_{estimate['quote_id']}_{datetime.today().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

